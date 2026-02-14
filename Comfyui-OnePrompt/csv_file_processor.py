

import os
import csv
from typing import Dict, List, Tuple

class csv_read:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "csv_path": ("STRING", {"default": "csv_file_path.csv"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("data",)
    FUNCTION = "csv_read"
    CATEGORY = "OnePrompt"

    @classmethod
    def IS_CHANGED(cls, **kwargs):
        return float("NaN")

    def csv_read(self, csv_path, row, column):
        # 调试日志
        print(f"=== CSV读取调试 ===")
        print(f"文件路径：{csv_path}")
        print(f"目标行：{row}，目标列：{column}")
        print(f"文件存在：{os.path.exists(csv_path)}")
        
        target_row = max(1, row)
        target_col = max(1, column)

        # 检查文件存在性
        if not os.path.exists(csv_path):
            err = f"Error: 文件不存在 → {csv_path}"
            print(err)
            return (err,)

        # 检查读权限
        if not os.access(csv_path, os.R_OK):
            err = f"Error: 无文件读取权限 → {csv_path}"
            print(err)
            return (err,)

        rows = []
        encodings = ["utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1"]
        read_success = False

        # 尝试多种编码+自动检测分隔符
        for encoding in encodings:
            try:
                with open(csv_path, mode="r", encoding=encoding, newline="") as file:
                    # 自动检测分隔符（处理逗号/制表符/分号）
                    sample = file.read(1024)
                    file.seek(0)
                    if csv.Sniffer().has_header(sample):
                        dialect = csv.Sniffer().sniff(sample)
                    else:
                        dialect = csv.excel  # 默认逗号分隔
                    reader = csv.reader(file, dialect=dialect)
                    rows = [r for r in reader if r]  # 过滤空行
                read_success = True
                print(f"成功读取，使用编码：{encoding}，分隔符：{dialect.delimiter}")
                break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                err = f"Error: 编码{encoding}读取失败 → {str(e)}"
                print(err)
                return (err,)

        # 检查读取结果
        if not read_success:
            err = "Error: 所有编码都无法读取文件（utf-8/utf-8-sig/gbk/gb2312/latin-1）"
            print(err)
            return (err,)

        if not rows:
            err = "Error: CSV文件为空（无任何行）"
            print(err)
            return (err,)

        # 检查行是否存在
        if target_row > len(rows):
            err = f"Error: 目标行{target_row}不存在 → 文件共{len(rows)}行"
            print(err)
            return (err,)

        target_row_data = rows[target_row - 1]
        # 检查列是否存在
        if target_col > len(target_row_data):
            err = f"Error: 目标列{target_col}不存在 → 第{target_row}行共{len(target_row_data)}列"
            print(err)
            return (err,)

        # 获取单元格值
        cell_value = target_row_data[target_col - 1]
        result = str(cell_value).strip() if cell_value is not None else ""
        print(f"读取成功 → 单元格值：{result}")
        
        return (result,)


class csv_write_data:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "csv_path": ("STRING", {"default": "csv_file_path.csv"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
                "data": ("STRING", {"forceInput": True}),
            },
            "optional": {}
        }
    
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_data"
    CATEGORY = "OnePrompt"
    OUTPUT_NODE = True  # 标记为输出节点，需返回UI数据
    
    @classmethod
    def IS_CHANGED(cls, **kwargs):  # 补充kwargs参数
        # 修复缓存问题：参数变化时重新执行（替代NaN）
        import hashlib
        m = hashlib.md5()
        for k, v in sorted(kwargs.items()):
            m.update(str(v).encode())
        return m.hexdigest()

    def write_data(self, csv_path, row, column, data):
        try:
            target_row = max(1, row)
            target_col = max(1, column)
            rows = []
            
            # 1. 读取现有CSV（兼容多种编码）
            encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312"]
            read_success = False
            if os.path.exists(csv_path):
                for encoding in encodings:
                    try:
                        with open(csv_path, 'r', encoding=encoding, newline='') as file:
                            reader = csv.reader(file)
                            rows = list(reader)
                        read_success = True
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        error_msg = f"读取文件失败: {str(e)}"
                        return {"ui": {"text": [error_msg]}, "result": (error_msg,)}
            
            # 2. 确保行列数足够（补空行/空列）
            # 补空行
            while len(rows) < target_row:
                rows.append([])
            # 补空列
            target_row_data = rows[target_row - 1]
            while len(target_row_data) < target_col:
                target_row_data.append("")
            
            # 3. 写入数据（过滤空字符串）
            if data.strip():
                rows[target_row - 1][target_col - 1] = data.strip()
                # 4. 保存文件（用utf-8-sig确保中文正常）
                with open(csv_path, 'w', newline='', encoding='utf-8-sig') as file:
                    writer = csv.writer(file)
                    writer.writerows(rows)
                success_msg = f"✅ 成功将数据写入：\n文件路径：{csv_path}\n行 {target_row} 列 {target_col}\n内容：{data.strip()}"
            else:
                success_msg = "⚠️ 未写入数据：输入内容为空"
            
            # 5. 输出节点返回UI数据（兼容ComfyUI显示）
            return {"ui": {"text": [success_msg]}, "result": (success_msg,)}
        
        # 针对性异常捕获
        except PermissionError as pe:
            error_msg = f"❌ 权限错误：无法写入文件\n原因：{str(pe)}\n路径：{csv_path}"
            return {"ui": {"text": [error_msg]}, "result": (error_msg,)}
        except FileNotFoundError as fnfe:
            error_msg = f"❌ 路径错误：文件/目录不存在\n原因：{str(fnfe)}\n路径：{csv_path}"
            return {"ui": {"text": [error_msg]}, "result": (error_msg,)}
        except Exception as e:
            error_msg = f"❌ 写入失败：{str(e)}"
            return {"ui": {"text": [error_msg]}, "result": (error_msg,)}



class csv_Std_Processor:
    CATEGORY = "OnePrompt"
    DESCRIPTION = "处理Excel和CSV文件：将xlsx转换为csv，或标准化csv文件"

    @classmethod
    def INPUT_TYPES(s) -> dict:
        return {
            "required": {
                "input_file": ("STRING", {"multiline": False, "default": ""}),
                "output_file": ("STRING", {"multiline": False, "default": ""}),
            }
        }

    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("output_path", "status")
    FUNCTION = "process_file"

    def process_file(self, input_file: str, output_file: str) -> Tuple[str, str]:
        if not input_file:
            return "", "Error: Input file path is required"

        input_file = input_file.strip('"')

        if not os.path.exists(input_file):
            return "", f"Error: Input file not found: {input_file}"

        if not output_file:
            base_name = os.path.splitext(input_file)[0]
            output_file = f"{base_name}_processed.csv"
        else:
            output_file = output_file.strip('"')

        file_ext = os.path.splitext(input_file)[1].lower()

        try:
            if file_ext == ".xlsx":
                self._convert_xlsx_to_csv(input_file, output_file)
            elif file_ext == ".csv":
                self._standardize_csv(input_file, output_file)
            else:
                return "", f"Error: Unsupported file type: {file_ext}"

            if os.path.exists(output_file):
                return output_file, "Success: File processed successfully"
            else:
                return "", "Error: Failed to create output file"
        except Exception as e:
            return "", f"Error: {str(e)}"

    def _convert_xlsx_to_csv(self, input_file: str, output_file: str):
        try:
            import pandas as pd
            df = pd.read_excel(input_file)
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
        except ImportError:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(input_file)
                ws = wb.active

                with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
            except ImportError:
                raise ImportError("Please install pandas or openpyxl for Excel processing")

    def _standardize_csv(self, input_file: str, output_file: str):
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'ansi', 'latin-1']
        rows = []

        for encoding in encodings:
            try:
                with open(input_file, 'r', encoding=encoding, newline='') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                break
            except Exception:
                continue

        if not rows:
            raise ValueError("Could not read CSV file with any encoding")

        with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerows(rows)




