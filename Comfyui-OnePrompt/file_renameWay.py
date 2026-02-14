from typing import List, Dict, Any
import os
import csv
import json
from typing import Tuple
import shutil
from pathlib import Path


class AnyType(str):
    def __eq__(self, _) -> bool:
        return True
    def __ne__(self, __value: object) -> bool:
        return False
ANY_TYPE = AnyType("*")


try:
    from openpyxl import load_workbook
    has_openpyxl = True
except ImportError:
    has_openpyxl = False


class file_CSVMapReader:
    CATEGORY = "OnePrompt"
    DESCRIPTION = "Read mapping information from CSV file, extracting column names and their corresponding unique IDs"

    @classmethod
    def INPUT_TYPES(s) -> dict:
        return {
            "required": {
                "csv_file_path": ("STRING", {"multiline": False, "default": ""}),
                "name_column": ("INT", {"default": 1, "min": 1, "max": 100, "step": 1}),
                "id_column": ("INT", {"default": 2, "min": 1, "max": 100, "step": 1}),
            }
        }

    RETURN_TYPES = ("LIST", "STRING")
    RETURN_NAMES = ("mapping_list", "status")
    FUNCTION = "read_csv_map"

    def read_csv_map(self, csv_file_path: str, name_column: int, id_column: int) -> tuple:
        try:
            csv_file_path = csv_file_path.strip('"').strip("'")
            csv_file_path = csv_file_path.replace('/', '\\')
            csv_file_path = os.path.normpath(csv_file_path)

            if not os.path.exists(csv_file_path):
                return ([], f"Error: CSV file not found at {csv_file_path}")

            if not csv_file_path.lower().endswith('.csv'):
                return ([], "Error: File is not a CSV file")

            mapping_list = []
            id_set = set()

            with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
                rows = list(csv.reader(csvfile))

                if len(rows) < 2:
                    return ([], "Error: CSV file is empty or has no data rows")

                header = rows[0]
                name_col_idx = name_column - 1
                if name_col_idx < 0 or name_col_idx >= len(header):
                    return ([], f"Error: Name column index {name_column} is out of range")

                id_col_idx = id_column - 1
                if id_col_idx < 0 or id_col_idx >= len(header):
                    return ([], f"Error: ID column index {id_column} is out of range")

                processed_rows = 0
                skipped_rows = 0
                empty_id_rows = 0

                for i, row in enumerate(rows[1:], start=2):
                    if len(row) <= max(name_col_idx, id_col_idx):
                        skipped_rows += 1
                        continue

                    try:
                        name = row[name_col_idx].strip()
                        id_value = row[id_col_idx].strip()

                        if not name:
                            skipped_rows += 1
                            continue

                        if not id_value:
                            empty_id_rows += 1
                            continue

                        if id_value in id_set:
                            return ([], f"Error: Duplicate ID found: {id_value} at row {i}")

                        id_set.add(id_value)
                        mapping_list.append({"id": id_value, "name": name})
                        processed_rows += 1
                    except Exception as e:
                        skipped_rows += 1
                        continue

            if not mapping_list:
                if empty_id_rows > 0:
                    if empty_id_rows == len(rows) - 1:
                        return ([], f"Warning: No valid mapping found in CSV. All {empty_id_rows} rows had empty ID values in column {id_column}. Please check your CSV file structure.")
                    else:
                        return ([], f"Warning: No valid mapping found in CSV. {empty_id_rows} rows had empty ID values. Please check if you're using the correct column index.")
                else:
                    return ([], f"Warning: No valid mapping found in CSV. Processed {processed_rows} rows, skipped {skipped_rows} rows")

            return (mapping_list, f"Success: Read {len(mapping_list)} mappings from CSV")

        except Exception as e:
            return ([], f"Error: {str(e)}")


class file_XLSXMapReader:
    CATEGORY = "OnePrompt"
    DESCRIPTION = "Read mapping information from XLSX file, extracting column names and their corresponding unique IDs"

    @classmethod
    def INPUT_TYPES(s) -> dict:
        return {
            "required": {
                "xlsx_file_path": ("STRING", {"multiline": False, "default": ""}),
                "name_column": ("INT", {"default": 1, "min": 1, "max": 100, "step": 1}),
                "id_column": ("INT", {"default": 2, "min": 1, "max": 100, "step": 1}),
                "sheet_name": ("STRING", {"multiline": False, "default": "Sheet1"}),
            }
        }

    RETURN_TYPES = ("LIST", "STRING")
    RETURN_NAMES = ("mapping_list", "status")
    FUNCTION = "read_xlsx_map"

    def read_xlsx_map(self, xlsx_file_path: str, name_column: int, id_column: int, sheet_name: str = "Sheet1") -> tuple:
        try:
            if not has_openpyxl:
                return ([], "Error: openpyxl is not installed. Please install it with 'pip install openpyxl'")

            xlsx_file_path = xlsx_file_path.strip('"').strip("'")
            xlsx_file_path = xlsx_file_path.replace('/', '\\')
            xlsx_file_path = os.path.normpath(xlsx_file_path)

            if not os.path.exists(xlsx_file_path):
                return ([], f"Error: XLSX file not found at {xlsx_file_path}")

            if not xlsx_file_path.lower().endswith('.xlsx'):
                return ([], "Error: File is not a XLSX file")

            mapping_list = []
            id_set = set()

            workbook = load_workbook(filename=xlsx_file_path, read_only=True)

            if sheet_name not in workbook.sheetnames:
                return ([], f"Error: Sheet '{sheet_name}' not found in XLSX file")

            sheet = workbook[sheet_name]

            name_col_idx = name_column - 1
            id_col_idx = id_column - 1

            processed_rows = 0
            skipped_rows = 0
            empty_id_rows = 0

            header_skipped = False

            for row in sheet.iter_rows(values_only=True):
                if not row or all(cell is None for cell in row):
                    continue

                if not header_skipped:
                    header_skipped = True
                    continue

                if len(row) <= max(name_col_idx, id_col_idx):
                    skipped_rows += 1
                    continue

                try:
                    name_cell = row[name_col_idx]
                    id_cell = row[id_col_idx]

                    name = str(name_cell).strip() if name_cell is not None else ""
                    id_value = str(id_cell).strip() if id_cell is not None else ""

                    if not name:
                        skipped_rows += 1
                        continue

                    if not id_value:
                        empty_id_rows += 1
                        continue

                    if id_value in id_set:
                        return ([], f"Error: Duplicate ID found: {id_value}")

                    id_set.add(id_value)
                    mapping_list.append({"id": id_value, "name": name})
                    processed_rows += 1
                except Exception as e:
                    skipped_rows += 1
                    continue

            workbook.close()

            if not mapping_list:
                if empty_id_rows > 0:
                    return ([], f"Warning: No valid mapping found in XLSX. {empty_id_rows} rows had empty ID values. Please check if you're using the correct column index.")
                else:
                    return ([], f"Warning: No valid mapping found in XLSX. Processed {processed_rows} rows, skipped {skipped_rows} rows")

            return (mapping_list, f"Success: Read {len(mapping_list)} mappings from XLSX")

        except Exception as e:
            return ([], f"Error: {str(e)}")


class file_MapRenamer:
    CATEGORY = "OnePrompt"
    DESCRIPTION = "Rename files in directory based on mapping information, matching filenames with names in mapping and renaming to corresponding IDs"

    @classmethod
    def INPUT_TYPES(s) -> dict:
        return {
            "required": {
                "mapping_list": ("LIST", {}),
                "directory_path": ("STRING", {"multiline": False, "default": ""}),
            },
            "optional": {
                "file_extensions": ("STRING", {"multiline": False, "default": "jpg|png|txt"}),
                "try_run": ("BOOLEAN", {"default": False}),
            }
        }

    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("rename_results", "status")
    FUNCTION = "rename_files"

    def rename_files(self, mapping_list: List[Dict[str, str]], directory_path: str,
                     file_extensions: str = "jpg|png|txt", try_run: bool = False) -> tuple:
        try:
            directory_path = directory_path.strip('"').strip("'")
            directory_path = directory_path.replace('/', '\\')
            directory_path = os.path.normpath(directory_path)

            if not os.path.exists(directory_path):
                return ("", f"Error: Directory not found at {directory_path}")

            if not os.path.isdir(directory_path):
                return ("", f"Error: {directory_path} is not a directory")

            if not isinstance(mapping_list, list) or not mapping_list:
                return ("", "Error: Invalid or empty mapping list")

            if '|' in file_extensions:
                extensions = [ext.strip() for ext in file_extensions.split('|')]
            else:
                extensions = [ext.strip() for ext in file_extensions.split(',')]

            normalized_extensions = []
            for ext in extensions:
                ext = ext.strip()
                if ext:
                    if ext.startswith('*'):
                        ext = ext[1:]
                    if not ext.startswith('.'):
                        ext = '.' + ext
                    normalized_extensions.append(ext.lower())

            name_to_id = {}
            for mapping in mapping_list:
                if isinstance(mapping, dict) and "name" in mapping and "id" in mapping:
                    name = mapping["name"].strip()
                    id_value = mapping["id"].strip()
                    if name and id_value:
                        name_to_id[name] = id_value

            if not name_to_id:
                return ("", "Error: No valid mappings found in mapping list")

            rename_operations = []
            renamed_count = 0
            skipped_count = 0

            for filename in os.listdir(directory_path):
                file_path = os.path.join(directory_path, filename)

                if not os.path.isfile(file_path):
                    continue

                ext_match = False
                file_ext = os.path.splitext(filename)[1].lower()

                if not normalized_extensions:
                    ext_match = True
                else:
                    for ext in normalized_extensions:
                        if file_ext == ext:
                            ext_match = True
                            break

                if not ext_match:
                    continue

                name_without_ext = os.path.splitext(filename)[0]
                ext = os.path.splitext(filename)[1]

                if name_without_ext in name_to_id:
                    new_id = name_to_id[name_without_ext]
                    new_filename = f"{new_id}{ext}"
                    new_file_path = os.path.join(directory_path, new_filename)

                    if os.path.exists(new_file_path):
                        skipped_count += 1
                        rename_operations.append(f"Skipped: {filename} -> {new_filename} (already exists)")
                        continue

                    if not try_run:
                        os.rename(file_path, new_file_path)
                        renamed_count += 1
                        rename_operations.append(f"Renamed: {filename} -> {new_filename}")
                    else:
                        rename_operations.append(f"Would rename: {filename} -> {new_filename}")
                        renamed_count += 1
                else:
                    skipped_count += 1
                    rename_operations.append(f"Skipped: {filename} (no matching mapping)")

            if try_run:
                status = f"Try run completed: Would rename {renamed_count} files, skipped {skipped_count} files"
            else:
                status = f"Rename completed: Renamed {renamed_count} files, skipped {skipped_count} files"

            results = "\n".join(rename_operations)

            return (results, status)

        except Exception as e:
            return ("", f"Error: {str(e)}")


class file_Rename_ByList:
    CATEGORY = "OnePrompt"
    DESCRIPTION = "Copy files from input list to target directory with new names (one-to-one mapping)"

    @classmethod
    def INPUT_TYPES(s) -> dict:
        return {
            "required": {
                "file_paths": (ANY_TYPE, {}),
                "name_list": (ANY_TYPE, {}),
                "target_directory": ("STRING", {"multiline": False, "default": ""}),
            },
            "optional": {
                "keep_original_ext": ("BOOLEAN", {"default": True}),
                "overwrite": ("BOOLEAN", {"default": False}),
                "try_run": ("BOOLEAN", {"default": False}),
            }
        }

    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("copy_results", "status")
    FUNCTION = "copy_rename_files"
    INPUT_IS_LIST = [True, True, False, False, False, False]

    def copy_rename_files(self,
                         file_paths: list,
                         name_list: list,
                         target_directory: str,
                         keep_original_ext: bool = True,
                         overwrite: bool = False,
                         try_run: bool = False) -> tuple:
        try:
            if isinstance(target_directory, list):
                target_directory = target_directory[0] if target_directory else ""

            if isinstance(target_directory, str):
                target_directory = target_directory.strip('"').strip("'").replace('\\', '/')
                target_directory = os.path.normpath(target_directory)
            else:
                return ("", "Error: Invalid target directory path")

            if not isinstance(file_paths, list) or len(file_paths) == 0:
                return ("", "Error: Invalid or empty file path list")
            if not isinstance(name_list, list) or len(name_list) == 0:
                return ("", "Error: Invalid or empty name list")

            if len(file_paths) != len(name_list):
                return ("", f"Error: File count ({len(file_paths)}) != Name count ({len(name_list)})")

            if not os.path.exists(target_directory):
                os.makedirs(target_directory, exist_ok=True)
            elif not os.path.isdir(target_directory):
                return ("", f"Error: {target_directory} is not a directory")

            copy_operations = []
            copied_count = 0
            skipped_count = 0

            if isinstance(try_run, list):
                try_run = try_run[0] if try_run else False
            try_run = bool(try_run)

            for idx, (src_file, new_name) in enumerate(zip(file_paths, name_list)):
                if isinstance(src_file, list):
                    src_file = src_file[0] if src_file else ""
                src_file = src_file.strip('"').strip("'") if isinstance(src_file, str) else str(src_file)

                if not os.path.isabs(src_file):
                    src_file = os.path.join(os.getcwd(), src_file)
                src_file = os.path.normpath(src_file)

                if not os.path.exists(src_file) or not os.path.isfile(src_file):
                    skipped_count += 1
                    copy_operations.append(f"Skipped [{idx+1}]: Source file not found - {src_file}")
                    continue

                original_name = os.path.basename(src_file)
                original_ext = os.path.splitext(original_name)[1] if keep_original_ext else ""
                if keep_original_ext and os.path.splitext(new_name)[1]:
                    new_name = os.path.splitext(new_name)[0]
                final_new_name = f"{new_name}{original_ext}"
                dest_file = os.path.join(target_directory, final_new_name)

                if os.path.exists(dest_file):
                    if overwrite:
                        action = "Overwritten"
                    else:
                        skipped_count += 1
                        copy_operations.append(f"Skipped [{idx+1}]: Target exists - {dest_file}")
                        continue
                else:
                    action = "Copied"

                if not try_run:
                    try:
                        shutil.copy2(src_file, dest_file)
                        copied_count += 1
                        copy_operations.append(f"{action} [{idx+1}]: {original_name} -> {final_new_name}")
                    except Exception as e:
                        skipped_count += 1
                        copy_operations.append(f"Failed [{idx+1}]: {original_name} -> {final_new_name} (Error: {str(e)})")
                else:
                    copied_count += 1
                    copy_operations.append(f"Would {action.lower()} [{idx+1}]: {original_name} -> {final_new_name}")

            status = f"{'Try run' if try_run else 'Operation'} completed: {'Would copy' if try_run else 'Copied'} {copied_count} files, skipped {skipped_count} files"
            results = "\n".join(copy_operations)
            return (results, status)

        except Exception as e:
            return ("", f"Error: {str(e)}")


class file_Organizer:
    CATEGORY = "OnePrompt"
    DESCRIPTION = "Organize and categorize files into folders based on file type extensions"

    FILE_TYPE_MAPPING = {
        '文档': ['.txt', '.doc', '.docx', '.pdf', '.xls', '.xlsx', '.ppt', '.pptx', '.md', '.csv', '.rtf'],
        '图片': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.svg', '.webp', '.ico'],
        '视频': ['.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.rmvb', '.mpeg'],
        '音频': ['.mp3', '.wav', '.flac', '.m4a', '.wma', '.ogg', '.aac'],
        '压缩包': ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2', '.iso'],
        '安装包': ['.exe', '.msi', '.dmg', '.pkg', '.apk'],
        '编程文件': ['.py', '.java', '.c', '.cpp', '.h', '.js', '.html', '.css', '.json', '.xml'],
        '其他文件': []
    }

    @classmethod
    def INPUT_TYPES(s) -> dict:
        return {
            "required": {
                "directory_path": ("STRING", {"multiline": False, "default": ""}),
            },
            "optional": {
                "try_run": ("BOOLEAN", {"default": False}),
                "create_folders": ("BOOLEAN", {"default": True}),
            }
        }

    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("organize_results", "status")
    FUNCTION = "organize_files"

    def organize_files(self, directory_path: str, try_run: bool = False, create_folders: bool = True) -> tuple:
        try:
            directory_path = directory_path.strip('"').strip("'")
            directory_path = directory_path.replace('/', '\\')
            directory_path = os.path.normpath(directory_path)

            if not os.path.exists(directory_path):
                return ("", f"Error: Directory not found at {directory_path}")

            if not os.path.isdir(directory_path):
                return ("", f"Error: {directory_path} is not a directory")

            organize_operations = []
            moved_count = 0
            skipped_count = 0
            folder_count = 0

            for file_name in os.listdir(directory_path):
                file_path = os.path.join(directory_path, file_name)

                if os.path.isdir(file_path):
                    continue

                file_suffix = Path(file_path).suffix.lower()
                category = None

                for cat, suffix_list in self.FILE_TYPE_MAPPING.items():
                    if file_suffix in suffix_list:
                        category = cat
                        break

                if not category:
                    category = '其他文件'

                target_folder = os.path.join(directory_path, category)

                if create_folders and try_run:
                    if not os.path.exists(target_folder):
                        folder_count += 1
                        organize_operations.append(f"Would create folder: {category}")

                if create_folders and not os.path.exists(target_folder):
                    os.makedirs(target_folder)
                    folder_count += 1

                target_file_path = os.path.join(target_folder, file_name)

                if os.path.exists(target_file_path):
                    file_stem = Path(file_name).stem
                    counter = 1
                    while os.path.exists(target_file_path):
                        new_file_name = f"{file_stem}_{counter}{file_suffix}"
                        target_file_path = os.path.join(target_folder, new_file_name)
                        counter += 1

                if not try_run:
                    try:
                        shutil.move(file_path, target_file_path)
                        moved_count += 1
                        organize_operations.append(f"Moved: {file_name} -> {category}")
                    except Exception as e:
                        skipped_count += 1
                        organize_operations.append(f"Failed: {file_name} -> {category} (Error: {str(e)})")
                else:
                    moved_count += 1
                    target_name = os.path.basename(target_file_path)
                    organize_operations.append(f"Would move: {file_name} -> {category}/{target_name}")

            if try_run:
                status = f"Try run completed: Would move {moved_count} files, create {folder_count} folders, skipped {skipped_count} files"
            else:
                status = f"Organize completed: Moved {moved_count} files, created {folder_count} folders, skipped {skipped_count} files"

            results = "\n".join(organize_operations)

            return (results, status)

        except Exception as e:
            return ("", f"Error: {str(e)}")





















